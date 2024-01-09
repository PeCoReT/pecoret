import io
from openpyxl import Workbook
from pecoret.reporting.generators import ExcelGenerator as BaseExcelGenerator


class ExcelGenerator(BaseExcelGenerator):
    def generate(self, entry):
        workbook = Workbook()
        ws = workbook.active
        ws.title = 'Findings'
        ws['A1'] = 'ID'
        ws['B1'] = 'Name'
        ws['C1'] = 'Vulnerability'
        ws['D1'] = 'Severity'
        ws['E1'] = 'Component'
        ws['F1'] = 'Description'
        ws['G1'] = 'Recommendation'
        ws['H1'] = 'Status'

        findings = self.context['findings']
        for finding in findings:
            if finding.recommendation:
                recommendation = finding.recommendation
            else:
                recommendation = finding.vulnerability.recommendation
            ws.append([
                finding.unique_id,
                finding.name,
                finding.vulnerability.name,
                finding.get_severity_display(),
                str(finding.component),
                finding.vulnerability.description,
                recommendation,
                finding.get_status_display()
            ])
        f = io.BytesIO()
        workbook.save(f)
        return f.getvalue(), self.content_type, self.file_extension
